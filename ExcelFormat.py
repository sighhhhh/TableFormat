import string
import os
import openpyxl
import configparser
from openpyxl.workbook import Workbook

script_dir = os.path.dirname(os.path.abspath(__file__))
config_file_path = os.path.join(script_dir,"Directory.ini")
config = configparser.ConfigParser()
config.read(config_file_path)


# 遍历多个excel文件
def read_dir(directory: string):
    filename = []
    filenames = []
    print(directory)
    for parents, dirnames, filenames in os.walk(directory):
        print(directory, ' & ', filenames, '&', len(filenames))
    for fileindex in range(len(filenames)):
        if ".xlsx" in filenames[fileindex]:
            filename.append(directory + filenames[fileindex])
    return filename


# 遍历读取表格的每一个 sheet 表
def read_table(files):
    list_temp = []
    for fileindex in files:
        book = openpyxl.load_workbook(fileindex)
        sheetNamelist = book.sheetnames
        # 遍历~一张 sheet 表
        for name in sheetNamelist:
            sheet = book[name]
            get_non_empty_cell_value(sheet,list_temp)
    return list_temp

def get_non_empty_cell_value(sheet,list_temp):
    # 遍历读取每一个 sheet 表中的 每一列
    for col in sheet.iter_cols():
        # 若整列中含不为空的数据项，则保留数据
        if any(cell.value is not None for cell in col):
            for cell in col:
                if cell.value is not None:
                    list_temp.append(cell.value)


# 填入临时表中
def write_table(listtemp: list):
    WorkTable = Workbook()
    sheet = WorkTable.active
    sheet.title = 'Sheet1'
    for raw in range(len(listtemp)):
        sheet.cell(raw + 1, 1, listtemp[raw])
    WorkTable.save(config.get("Directory","OutPath"))


if __name__ == "__main__":
    directory = config.get("Directory", "ReadPath")
    files = read_dir(directory)
    write_table(read_table(files))